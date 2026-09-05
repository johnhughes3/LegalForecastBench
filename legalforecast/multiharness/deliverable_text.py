"""Deterministic visible-text extraction from sealed LAB deliverables.

The Harvey LAB bridge carries DOCX and XLSX OOXML containers rather than plain
text. A paid judge has to be shown the candidate's actual work product, so
those bytes must be turned into deterministic text before they reach a
provider.

This module is therefore *judge-input-determining* code: what it returns is
what the judge grades, so every ambiguous case fails closed rather than
returning a partial document.  A silently truncated or silently empty
extraction would produce a confident, billed verdict about text nobody sent,
which is the failure mode the per-criterion seam exists to prevent.

Provider-free and deterministic: no network, no credentials, and the same
bytes always yield the same string.
"""

from __future__ import annotations

import io
import zipfile
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from xml.parsers import expat

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# The main document part every WordprocessingML package must contain.
DOCX_DOCUMENT_PART = "word/document.xml"
# Parts that carry body text this extractor does not render. The document part
# references a footnote by ID only, so footnote prose never appears in it: a memo
# arguing in its footnotes would otherwise be graded without that argument.
DOCX_UNRENDERED_TEXT_PARTS = ("word/footnotes.xml", "word/endnotes.xml")
_NAMESPACE_SEPARATOR = "|"
_WORDPROCESSING_NAMESPACE = (
    "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
)

# A LAB memo is a text document; anything near this size is not one. The cap is
# applied to the *declared* and the *decompressed* size, so a zip bomb is
# refused before it is expanded rather than after.
DOCX_MAX_DOCUMENT_PART_BYTES = 8 * 1024 * 1024
XLSX_MAX_WORKSHEETS = 64
XLSX_MAX_CELL_SLOTS = 250_000
XLSX_MAX_RENDERED_CHARS = 8 * 1024 * 1024


class DeliverableTextError(ValueError):
    """Raised when deliverable text cannot be extracted fail-closed."""


def deliverable_visible_text(payload: bytes, *, basename: str) -> str:
    """Extract one supported deliverable using its authenticated basename."""

    if basename.endswith(".docx"):
        return docx_visible_text(payload)
    if basename.endswith(".xlsx"):
        return xlsx_visible_text(payload)
    raise DeliverableTextError("deliverable basename has an unsupported suffix")


def _qualified(tag: str) -> str:
    return f"{_WORDPROCESSING_NAMESPACE}{_NAMESPACE_SEPARATOR}{tag}"


def docx_visible_text(
    payload: bytes,
    *,
    max_document_part_bytes: int = DOCX_MAX_DOCUMENT_PART_BYTES,
) -> str:
    """Return the visible text of a ``.docx`` payload in document order.

    Only genuinely visible runs are returned: ``w:t`` text, ``w:tab`` tabs, and
    ``w:br``/``w:cr`` breaks, with each ``w:p`` starting a new line. Field
    instructions (``w:instrText``) and tracked deletions (``w:delText``) are
    excluded because neither is part of the delivered work product. Table text
    is included, since table cells are built from ordinary paragraphs.

    Raises ``DeliverableTextError`` for anything that would otherwise yield a
    partial or empty rendering.
    """

    if type(payload) is not bytes or not payload:
        raise DeliverableTextError("deliverable payload must be non-empty bytes")
    parts = _read_package_parts(payload, max_document_part_bytes)
    for name in DOCX_UNRENDERED_TEXT_PARTS:
        note_part = parts.get(name)
        if note_part is not None and _has_text(note_part):
            # Refuse rather than render around it. Extracting notes correctly
            # means deciding where they belong in reading order, which is a
            # design question; grading a memo without its footnotes is not a
            # question at all -- it is the silent-partial-input failure this
            # module exists to prevent.
            raise DeliverableTextError(
                "deliverable carries footnote or endnote text this extractor "
                "does not render"
            )
    text = _visible_text(parts[DOCX_DOCUMENT_PART])
    if not text:
        # A structurally valid file that renders to nothing must not be graded:
        # the judge would return a confident verdict about an empty document.
        raise DeliverableTextError("deliverable contains no extractable text")
    return text


def xlsx_visible_text(
    payload: bytes,
    *,
    max_worksheets: int = XLSX_MAX_WORKSHEETS,
    max_cell_slots: int = XLSX_MAX_CELL_SLOTS,
    max_rendered_chars: int = XLSX_MAX_RENDERED_CHARS,
) -> str:
    """Render worksheet names, coordinates, values, and formulas in order.

    Coordinates preserve sparse layout without synthesizing a rectangular
    CSV. Formulas are rendered as formulas (``data_only=False``), because a
    workbook's cached values are optional and may be stale. The rectangular
    worksheet dimensions and rendered output are bounded before the text can
    become judge input.
    """

    if type(payload) is not bytes or not payload:
        raise DeliverableTextError("deliverable payload must be non-empty bytes")
    for value, name in (
        (max_worksheets, "worksheet cap"),
        (max_cell_slots, "cell-slot cap"),
        (max_rendered_chars, "rendered-character cap"),
    ):
        if type(value) is not int or value <= 0:
            raise DeliverableTextError(f"{name} must be positive")
    try:
        workbook = load_workbook(
            io.BytesIO(payload),
            read_only=True,
            data_only=False,
            keep_links=False,
            rich_text=False,
        )
    except (
        OSError,
        ValueError,
        KeyError,
        zipfile.BadZipFile,
        InvalidFileException,
    ) as exc:
        raise DeliverableTextError(
            "deliverable is not a readable SpreadsheetML package"
        ) from exc
    try:
        worksheets = workbook.worksheets
        if not worksheets:
            raise DeliverableTextError("deliverable contains no worksheets")
        if len(worksheets) > max_worksheets:
            raise DeliverableTextError("deliverable contains too many worksheets")

        rendered: list[str] = []
        total_slots = 0
        total_chars = 0
        has_value = False
        for worksheet in worksheets:
            rows = max(worksheet.max_row, 1)
            columns = max(worksheet.max_column, 1)
            total_slots += rows * columns
            if total_slots > max_cell_slots:
                raise DeliverableTextError("deliverable contains too many cell slots")
            header = f"=== Sheet: {_escape_spreadsheet_text(worksheet.title)} ==="
            rendered.append(header)
            total_chars += len(header) + 1
            for row in worksheet.iter_rows():
                cells: list[str] = []
                for cell in row:
                    if cell.value is None:
                        continue
                    has_value = True
                    cells.append(
                        f"{cell.coordinate}={_render_spreadsheet_value(cell.value)}"
                    )
                if cells:
                    line = "\t".join(cells)
                    rendered.append(line)
                    total_chars += len(line) + 1
                    if total_chars > max_rendered_chars:
                        raise DeliverableTextError(
                            "deliverable renders beyond the character limit"
                        )
        if not has_value:
            raise DeliverableTextError(
                "deliverable contains no extractable cell values"
            )
        return "\n".join(rendered)
    finally:
        workbook.close()


def _render_spreadsheet_value(value: object) -> str:
    if isinstance(value, str):
        return _escape_spreadsheet_text(value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int | float | Decimal):
        return str(value)
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    raise DeliverableTextError("deliverable contains an unsupported cell value")


def _escape_spreadsheet_text(value: str) -> str:
    return value.replace("\\", "\\\\").replace("\t", "\\t").replace("\n", "\\n")


def _read_package_parts(
    payload: bytes, max_document_part_bytes: int
) -> dict[str, bytes]:
    """Read the document part, plus any note parts, under one byte cap each."""

    if type(max_document_part_bytes) is not int or max_document_part_bytes <= 0:
        raise DeliverableTextError("document part byte cap must be positive")
    parts: dict[str, bytes] = {}
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            names = set(archive.namelist())
            if DOCX_DOCUMENT_PART not in names:
                raise DeliverableTextError(
                    "deliverable is not a WordprocessingML package"
                )
            wanted = (DOCX_DOCUMENT_PART, *DOCX_UNRENDERED_TEXT_PARTS)
            for name in wanted:
                if name not in names:
                    continue
                info = archive.getinfo(name)
                if info.file_size > max_document_part_bytes:
                    raise DeliverableTextError("deliverable part is too large")
                with archive.open(info) as handle:
                    # Read one byte past the cap so a lying zip header is caught
                    # by the decompressed length rather than trusted at face value.
                    body = handle.read(max_document_part_bytes + 1)
                if len(body) > max_document_part_bytes:
                    raise DeliverableTextError("deliverable part is too large")
                parts[name] = body
    except (OSError, zipfile.BadZipFile, RuntimeError) as exc:
        raise DeliverableTextError("deliverable is not a readable zip package") from exc
    return parts


def _walk_text_events(document: bytes) -> list[tuple[str, str]]:
    """Stream (kind, text) events from a WordprocessingML part, DTD-free.

    Parsed with expat directly rather than ``ElementTree.fromstring`` so the
    DTD refusal is a parser callback rather than a byte scan of the source.
    A byte scan is not sound: expat honours the encoding declared in the XML
    prolog, so a part declared ``UTF-16`` hides the literal bytes of a
    ``<!DOCTYPE`` from any ASCII substring search while expat still parses and
    expands the entities it declares. Refusing at the callback rejects every
    encoding, which is what kills the entity-expansion class outright.
    """

    parser = expat.ParserCreate(namespace_separator=_NAMESPACE_SEPARATOR)
    # Entity expansion needs a DTD, and a WordprocessingML part never has one.
    parser.StartDoctypeDeclHandler = _refuse_doctype
    parser.EntityDeclHandler = _refuse_entity
    parser.ExternalEntityRefHandler = _refuse_external_entity

    events: list[tuple[str, str]] = []
    current_text: list[str] = []
    text_run = _qualified("t")
    paragraph = _qualified("p")
    tab = _qualified("tab")
    breaks = {_qualified("br"), _qualified("cr")}

    def start(name: str, _attributes: dict[str, str]) -> None:
        if name == paragraph:
            events.append(("paragraph", ""))
        elif name == tab:
            events.append(("text", "\t"))
        elif name in breaks:
            events.append(("text", "\n"))
        elif name == text_run:
            current_text.clear()

    def characters(data: str) -> None:
        # Expat may deliver one logical run in several callbacks.
        current_text.append(data)

    def end(name: str) -> None:
        if name == text_run:
            events.append(("run", "".join(current_text)))
            current_text.clear()

    parser.StartElementHandler = start
    parser.CharacterDataHandler = characters
    parser.EndElementHandler = end
    try:
        parser.Parse(document, True)
    except expat.ExpatError as exc:
        raise DeliverableTextError(
            "deliverable document part is not well-formed XML"
        ) from exc
    return events


def _refuse_doctype(*_args: object, **_kwargs: object) -> None:
    raise DeliverableTextError("deliverable document part declares a DTD")


def _refuse_entity(*_args: object, **_kwargs: object) -> None:
    raise DeliverableTextError("deliverable document part declares an entity")


def _refuse_external_entity(*_args: object, **_kwargs: object) -> bool:
    raise DeliverableTextError("deliverable document part references external data")


def _has_text(document: bytes) -> bool:
    """Report whether a part carries any non-whitespace text run.

    Word writes a footnotes part into every document containing only empty
    separator stubs, so presence of the part proves nothing; presence of actual
    ``w:t`` content does.
    """

    return any(
        text.strip() for kind, text in _walk_text_events(document) if kind == "run"
    )


def _visible_text(document: bytes) -> str:
    parts: list[str] = []
    started = False
    for kind, text in _walk_text_events(document):
        if kind == "paragraph":
            if started:
                parts.append("\n")
            started = True
        else:
            parts.append(text)
    rendered = "".join(parts)
    return "\n".join(line.rstrip() for line in rendered.split("\n")).strip()
